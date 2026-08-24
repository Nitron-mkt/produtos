#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Adapta a planilha de produtos Nitron 2026 para a Matriz de Etiquetas.

Le o modelo vazio (Nitron_Matriz_Etiquetas.xlsx) e a lista de produtos, e
escreve um ficheiro novo com os mesmos separadores, as mesmas convencoes de
cor (cinzento = Sankhya, amarelo = Marketing) e a Matriz_SKU preenchida.

Uso:
    python3 scripts/etiquetas/gerar_matriz.py
    python3 scripts/etiquetas/gerar_matriz.py --produtos outro.xlsx

Nao altera nenhum dos ficheiros de entrada.
"""
from __future__ import annotations

import argparse
import re
import sys
import unicodedata
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))
import regras as R  # noqa: E402

RAIZ = Path(__file__).resolve().parents[2]
MODELO = RAIZ / "etiquetas/fonte/Nitron_Matriz_Etiquetas_MODELO.xlsx"
PRODUTOS = RAIZ / "etiquetas/fonte/produtos-nitron-2026.csv"
SAIDA = RAIZ / "etiquetas/Nitron_Matriz_Etiquetas_PREENCHIDA.xlsx"
DIR_CSV = RAIZ / "etiquetas/csv"

# Convencoes visuais do modelo, lidas do proprio ficheiro modelo.
TURQUESA = "1B7F8C"
CINZA_FUNDO = "E8E8E8"
AMARELO = "FFFF00"
CINZA_TEXTO = "404040"

CAB_MATRIZ = ["CODPROD", "REFERENCIA", "DESCRICAO SANKHYA", "LAYOUT",
              "NOME PT", "NOME EN", "NOME ES", "CAP VALOR", "CAP UNID",
              "QTD PECAS", "LARG CM", "ALT CM", "PROF CM", "SELOS",
              "APLICACOES", "URL QR"]
# Colunas 1..3 vem do Sankhya (cinzento). 4..16 sao do Marketing (amarelo).
N_COLS_SANKHYA = 3


# ---------------------------------------------------------------------------
# utilitarios de texto
# ---------------------------------------------------------------------------
def sem_acento(s: str) -> str:
    return "".join(c for c in unicodedata.normalize("NFD", s)
                   if unicodedata.category(c) != "Mn")


def num(v):
    """'Ø22,5' -> (22.5, True) ; '18.' -> (18.0, False) ; '' -> (None, False)"""
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None, False
    s = str(v).strip()
    if not s or s == "-":
        return None, False
    diametro = "Ø" in s or "ø" in s
    s = s.replace("Ø", "").replace("ø", "").replace(",", ".").rstrip(".")
    s = re.sub(r"\.(?=\D|$)", "", s)
    try:
        return round(float(s), 2), diametro
    except ValueError:
        return None, diametro


CAP_RE = re.compile(r"^([\d.,]+)\s*(ml|l|kg|g)$", re.IGNORECASE)


def capacidade(vol):
    """'1,1L' -> (1.1, 'L', 1100.0, False). Terceiro valor e ml equivalente."""
    if vol is None or (isinstance(vol, float) and pd.isna(vol)):
        return None, None, None, False
    bruto = str(vol).strip()
    if not bruto or bruto == "-":
        return None, None, None, False
    multipla = bool(re.search(r"[\n/]", bruto))
    s = re.split(r"[\n/]", bruto)[0].strip().replace(",", ".")
    m = CAP_RE.match(s)
    if not m:
        return None, None, None, multipla
    valor = float(m.group(1))
    unid = m.group(2).upper()
    if unid == "ML":
        ml = valor
    elif unid == "L":
        ml = valor * 1000
    else:                       # KG e G sao peso, nao entram na faixa de volume
        ml = None
    if valor == int(valor):
        valor = int(valor)
    return valor, unid, ml, multipla


def limpar_nome(nome: str, cores: list[str]) -> str:
    """Remove codigo interno, capacidade e cor do fim do nome comercial."""
    if not isinstance(nome, str):
        return ""
    s = re.sub(r"\s*\([^)]*\)\s*$", "", nome).replace("\n", " ").strip()
    padrao_cap = r"\d+(?:[.,]\d+)?\s*(?:ml|mL|ML|l|L|litros?|kg|KG|Kg|g|G|gr)"
    # capacidade a meio do nome ("- 1L Cafe" -> "- Cafe"): a capacidade tem
    # coluna propria, no nome so rouba caracteres ao limite do layout
    s = re.sub(r"(\s-\s)" + padrao_cap + r"\s+", r"\1", s)
    anterior = None
    while anterior != s:
        anterior = s
        s = re.sub(padrao_cap + r"\s*$", "", s).strip()
        s = re.sub(r"[\s\-/]+$", "", s).strip()
        for cor in sorted(cores, key=len, reverse=True):
            s = re.sub(r"(?:\s*[-/]\s*|\s+e\s+|\s+y\s+|\s+and\s+|\s+)"
                       + re.escape(cor) + r"\s*$", "", s, flags=re.IGNORECASE).strip()
        s = re.sub(r"[\s\-/]+$", "", s).strip()
    return re.sub(r"\s{2,}", " ", s).strip(" -/")


def abreviar(nome: str, limite: int, tabela) -> tuple[str, list[str]]:
    """Aplica abreviaturas ate caber no limite. Devolve o nome e o que usou."""
    usadas = []
    s = nome
    for de, para in tabela:
        if len(s) <= limite:
            break
        if de.lower() in s.lower():
            s = re.sub(re.escape(de), para, s, flags=re.IGNORECASE)
            s = re.sub(r"\s{2,}", " ", s).strip(" -/")
            usadas.append(f"{de.strip()}>{para.strip() or '(removido)'}")
    return s, usadas


def ean_valido(ean: str) -> bool:
    s = re.sub(r"\D", "", str(ean or ""))
    if len(s) != 13:
        return False
    soma = sum(int(d) * (1 if i % 2 == 0 else 3) for i, d in enumerate(s[:12]))
    return (10 - soma % 10) % 10 == int(s[12])


# ---------------------------------------------------------------------------
# regras de conteudo
# ---------------------------------------------------------------------------
def escolher_layout(categoria, ml, larg, alt):
    if categoria in R.CATEGORIA_LAYOUT_FIXO:
        lay = R.CATEGORIA_LAYOUT_FIXO[categoria]
        return lay, f"categoria {categoria}"
    if ml is not None:
        for teto, lay in R.FAIXAS_CAPACIDADE:
            if ml <= teto:
                return lay, f"capacidade {ml:.0f} ml"
    area = (larg or 0) * (alt or 0)
    if area:
        for teto, lay in R.FAIXAS_AREA_FACE:
            if area <= teto:
                return lay, f"face {area:.0f} cm2"
    return 2, "sem capacidade nem dimensao (default)"


def deduzir_selos(categoria, texto, transparente, max_selos):
    evidencia = [cod for pad, cod in R.SELOS_POR_PALAVRA
                 if re.search(pad, texto, re.IGNORECASE)]
    if transparente:
        evidencia.append(R.SELO_TRANSPARENTE)
    base = R.SELOS_POR_CATEGORIA.get(categoria, [])
    todos = list(dict.fromkeys(evidencia + base + R.SELOS_UNIVERSAIS))
    ordenados = sorted(todos, key=lambda c: R.PRIORIDADE_SELOS.index(c)
                       if c in R.PRIORIDADE_SELOS else 99)
    return (ordenados[:max_selos], ordenados[max_selos:],
            [c for c in evidencia if c in ordenados])


def deduzir_aplicacoes(categoria, nome, texto, max_aplic):
    for padrao, codigos in R.APLICACOES_OVERRIDE_NOME:
        if re.search(padrao, nome.strip(), re.IGNORECASE):
            return codigos[:max_aplic], []
    evidencia = [cod for pad, cod in R.APLICACOES_POR_PALAVRA
                 if re.search(pad, texto, re.IGNORECASE)]
    base = R.APLICACOES_POR_CATEGORIA.get(categoria, [])
    todos = list(dict.fromkeys(evidencia + base))
    return todos[:max_aplic], todos[max_aplic:]


QTD_RE = re.compile(r"com\s+(\d+)\s*(peças|peca|pecas|potes|marmitas|"
                    r"molheiras|kits)", re.IGNORECASE)


def qtd_pecas(nome: str) -> int:
    m = QTD_RE.search(nome or "")
    if m:
        return int(m.group(1))
    m = re.match(r"\s*(?:kit|conjunto)\b.*?(\d+)\s*(?:peças|pecas|potes)",
                 nome or "", re.IGNORECASE)
    return int(m.group(1)) if m else 1


# ---------------------------------------------------------------------------
# construcao das linhas
# ---------------------------------------------------------------------------
def construir(df: pd.DataFrame):
    linhas, auditoria = [], []
    ean_vistos = {}
    for _, p in df.iterrows():
        ref = str(p["REF"]).strip()
        cat_bruta = str(p["CATEGORIA"]).strip()
        cat = R.CATEGORIA_CANONICA.get(cat_bruta, cat_bruta)
        nome_orig = str(p["PORTUGUÊS"]).strip()
        descricao = "" if pd.isna(p.get("DESCRIÇÃO")) else str(p["DESCRIÇÃO"])
        if descricao.strip() in ("#N/A", "#VALUE!"):
            descricao = ""

        cap_v, cap_u, ml, cap_multipla = capacidade(p.get("VOL"))
        larg, d1 = num(p.get("COM"))
        prof, d2 = num(p.get("LAR"))
        altu, d3 = num(p.get("ALT"))

        layout, regra_layout = escolher_layout(cat, ml, larg, altu)
        cfg = R.LAYOUTS[layout]
        limite = R.max_caracteres(layout)

        base_pt = limpar_nome(nome_orig, R.CORES_PT)
        base_en = limpar_nome(p.get("INGLES"), R.CORES_EN)
        base_es = limpar_nome(p.get("ESPANHOL"), R.CORES_ES)
        nome_pt, abv_pt = abreviar(base_pt, limite, R.ABREVIATURAS_PT)
        nome_en, _ = abreviar(base_en, limite, R.ABREVIATURAS_EN)
        nome_es, _ = abreviar(base_es, limite, R.ABREVIATURAS_ES)

        texto = sem_acento(f"{nome_orig} {descricao} {cat}").lower() + " " \
            + f"{nome_orig} {descricao} {cat}".lower()
        transparente = bool(re.search(r"transparen|cristal", nome_orig, re.I))
        selos, selos_cortados, selos_evid = deduzir_selos(
            cat, texto, transparente, cfg["max_selos"])
        aplic, aplic_cortadas = deduzir_aplicacoes(
            cat, nome_orig, texto, cfg["max_aplic"])

        linhas.append([
            None,                       # CODPROD - nao existe no ficheiro de origem
            ref,
            nome_orig,
            layout,
            nome_pt, nome_en, nome_es,
            cap_v, cap_u,
            qtd_pecas(nome_orig),
            larg, altu, prof,
            ",".join(selos),
            ",".join(aplic),
            R.BASE_URL_QR + ref,
        ])

        alertas = []
        if not base_en:
            alertas.append("SEM NOME EN")
        if not base_es:
            alertas.append("SEM NOME ES")
        if cap_v is None:
            alertas.append("SEM CAPACIDADE")
        if cap_multipla:
            alertas.append("CAPACIDADE MULTIPLA NA ORIGEM")
        if len(nome_pt) > limite:
            alertas.append(f"NOME PT EXCEDE LIMITE ({len(nome_pt)}>{limite})")
        if d1 or d2 or d3:
            alertas.append("DIMENSAO E DIAMETRO")
        if None in (larg, altu, prof):
            alertas.append("DIMENSAO EM FALTA")
        if not descricao:
            alertas.append("SEM DESCRICAO NA ORIGEM")
        ean = str(p.get("EAN") or "").strip()
        if not ean_valido(ean):
            alertas.append("EAN INVALIDO")
        if ean in ean_vistos:
            alertas.append(f"EAN DUPLICADO (tambem em {ean_vistos[ean]})")
        else:
            ean_vistos[ean] = ref
        if cat_bruta != cat:
            alertas.append(f"CATEGORIA NORMALIZADA ({cat_bruta})")

        auditoria.append([
            ref, ean, cat, layout, regra_layout,
            nome_orig, nome_pt, len(nome_pt), limite,
            "; ".join(abv_pt), ",".join(selos_evid), ",".join(selos_cortados),
            ",".join(aplic_cortadas), " | ".join(alertas),
        ])
    return linhas, auditoria


# ---------------------------------------------------------------------------
# escrita
# ---------------------------------------------------------------------------
def estilo_dados(cell, sankhya: bool):
    cell.font = Font(name="Arial", size=10)
    cell.fill = PatternFill("solid", fgColor=CINZA_FUNDO if sankhya else AMARELO)
    fina = Side(style="thin", color="BFBFBF")
    cell.border = Border(left=fina, right=fina, top=fina, bottom=fina)


def cabecalho(ws, titulo, subtitulo, colunas, larguras):
    ws["A1"] = titulo
    ws["A1"].font = Font(name="Arial", size=14, bold=True, color=TURQUESA)
    ws["A2"] = subtitulo
    ws["A2"].font = Font(name="Arial", size=10, color=CINZA_TEXTO)
    fina = Side(style="thin", color="BFBFBF")
    for i, nome in enumerate(colunas, 1):
        c = ws.cell(row=3, column=i, value=nome)
        c.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=TURQUESA)
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = Border(left=fina, right=fina, top=fina, bottom=fina)
        ws.column_dimensions[get_column_letter(i)].width = larguras[i - 1]
    ws.freeze_panes = "A4"


def preencher_matriz(ws, linhas):
    # o modelo traz a linha 4 como exemplo em cinzento italico e a 5 em branco
    ws.delete_rows(4, ws.max_row - 3)
    for r, valores in enumerate(linhas, start=4):
        for c, v in enumerate(valores, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            estilo_dados(cell, sankhya=c <= N_COLS_SANKHYA)
    ws.auto_filter.ref = f"A3:{get_column_letter(len(CAB_MATRIZ))}{3 + len(linhas)}"


def preencher_layouts(ws):
    for i, cod in enumerate(sorted(R.LAYOUTS), start=4):
        cfg = R.LAYOUTS[cod]
        ws.cell(row=i, column=3, value=cfg["largura_mm"])
        ws.cell(row=i, column=4, value=cfg["altura_mm"])
        ws.cell(row=i, column=5, value=R.max_caracteres(cod))
        ws.cell(row=i, column=6, value=cfg["max_selos"])
        ws.cell(row=i, column=7, value=cfg["max_aplic"])
        ws.cell(row=i, column=8, value=cfg["tem_qr"])
        for col in range(3, 9):
            estilo_dados(ws.cell(row=i, column=col), sankhya=False)


def acrescentar_aplicacoes_novas(ws):
    existentes = {ws.cell(row=r, column=1).value
                  for r in range(4, ws.max_row + 1)}
    ultimo = max((r for r in range(4, ws.max_row + 1)
                  if ws.cell(row=r, column=1).value), default=3)
    seq = max((int(str(ws.cell(row=r, column=5).value or "APLC0000")[4:])
               for r in range(4, ultimo + 1)), default=0)
    linha = ultimo + 1
    novas = 0
    for cod, pt, en, es in R.APLICACOES_NOVAS:
        if cod in existentes:
            continue
        seq += 1
        valores = [cod, pt, en, es, f"APLC{seq:04d}", "S", "S"]
        for c, v in enumerate(valores, start=1):
            cell = ws.cell(row=linha, column=c, value=v)
            estilo_dados(cell, sankhya=c <= 5)
        linha += 1
        novas += 1
    return novas


def aba_auditoria(wb, auditoria):
    ws = wb.create_sheet("Auditoria")
    cols = ["REFERENCIA", "EAN", "CATEGORIA", "LAYOUT", "PORQUE ESTE LAYOUT",
            "NOME NA ORIGEM", "NOME PT PROPOSTO", "CARACTERES", "LIMITE",
            "ABREVIATURAS APLICADAS", "SELOS COM EVIDENCIA NO TEXTO",
            "SELOS CORTADOS PELO LIMITE", "APLICACOES CORTADAS", "ALERTAS"]
    largs = [14, 15, 16, 8, 26, 42, 34, 11, 8, 30, 26, 24, 24, 52]
    cabecalho(ws, "AUDITORIA DO PREENCHIMENTO AUTOMATICO",
              "Uma linha por SKU. Serve para o Marketing conferir e corrigir. "
              "Nao e para importar no Sankhya.", cols, largs)
    fina = Side(style="thin", color="BFBFBF")
    vermelho = Font(name="Arial", size=10, color="C00000")
    for r, linha in enumerate(auditoria, start=4):
        for c, v in enumerate(linha, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = Font(name="Arial", size=10)
            cell.border = Border(left=fina, right=fina, top=fina, bottom=fina)
        if linha[-1]:
            ws.cell(row=r, column=14).font = vermelho
    ws.auto_filter.ref = f"A3:N{3 + len(auditoria)}"
    return ws


def aba_impressao(wb, df, linhas):
    """Dados da zona branca: o que as impressoras internas precisam."""
    ws = wb.create_sheet("Impressao_Interna")
    cols = ["REFERENCIA", "EAN", "LAYOUT", "NOME PT", "CAPACIDADE",
            "QTD PECAS", "SELOS", "APLICACOES", "URL QR"]
    largs = [14, 16, 8, 34, 14, 10, 26, 30, 26]
    cabecalho(ws, "DADOS PARA A IMPRESSAO INTERNA",
              "Um registo por SKU para o software da impressora de codigo de "
              "barras (Zebra, Argox, BarTender). O EAN nao consta da Matriz_SKU "
              "porque nao faz parte do conteudo grafico da etiqueta.", cols, largs)
    fina = Side(style="thin", color="BFBFBF")
    eans = dict(zip(df["REF"].astype(str).str.strip(),
                    df["EAN"].astype(str).str.strip()))
    for r, l in enumerate(linhas, start=4):
        cap = f"{l[7]} {l[8]}" if l[7] is not None else ""
        valores = [l[1], eans.get(l[1], ""), l[3], l[4], cap, l[9],
                   l[13], l[14], l[15]]
        for c, v in enumerate(valores, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = Font(name="Arial", size=10)
            cell.border = Border(left=fina, right=fina, top=fina, bottom=fina)
    ws.auto_filter.ref = f"A3:I{3 + len(linhas)}"
    return ws


def aba_regras(wb, linhas, n_matriz):
    ws = wb.create_sheet("Regras_Auto")
    ws["A1"] = "REGRAS DO PREENCHIMENTO AUTOMATICO"
    ws["A1"].font = Font(name="Arial", size=14, bold=True, color=TURQUESA)
    ws["A2"] = ("Tudo nesta folha e proposta do Marketing, nao vem do Sankhya. "
                "Para mudar criterio, editar scripts/etiquetas/regras.py e "
                "correr de novo o gerador.")
    ws["A2"].font = Font(name="Arial", size=10, color=CINZA_TEXTO)
    ws.column_dimensions["A"].width = 34
    for col in "BCDEFGH":
        ws.column_dimensions[col].width = 15
    ws.column_dimensions["I"].width = 46

    neg = Font(name="Arial", size=10, bold=True)
    normal = Font(name="Arial", size=10)

    def titulo(r, txt):
        ws.cell(row=r, column=1, value=txt).font = Font(
            name="Arial", size=11, bold=True, color=TURQUESA)

    # --- limite de caracteres, calculado com formulas
    titulo(4, "1. COMO SAI O MAX CARAC NOME")
    heads = ["LAYOUT", "JANELA LARG MM", "MARGEM MM", "CORPO PT",
             "LARG CARACTER MM", "LINHAS", "CABE NA JANELA",
             "TETO EDITORIAL", "MAX CARAC NOME"]
    for i, h in enumerate(heads, start=1):
        c = ws.cell(row=5, column=i, value=h)
        c.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=TURQUESA)
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    for i, cod in enumerate(sorted(R.LAYOUTS), start=6):
        cfg = R.LAYOUTS[cod]
        ws.cell(row=i, column=1, value=cod).font = normal
        ws.cell(row=i, column=2, value=cfg["janela_larg_mm"]).font = normal
        ws.cell(row=i, column=3, value=R.MARGEM_JANELA_MM).font = normal
        ws.cell(row=i, column=4, value=cfg["corpo_pt"]).font = normal
        f = ws.cell(row=i, column=5,
                    value=f"=D{i}*{R.MM_POR_PONTO}*{R.FATOR_CARACTERE}")
        f.font = normal
        f.number_format = "0.00"
        ws.cell(row=i, column=6, value=cfg["linhas"]).font = normal
        ws.cell(row=i, column=7,
                value=f"=INT((B{i}-2*C{i})/E{i})*F{i}").font = normal
        ws.cell(row=i, column=8, value=cfg["teto_editorial"]).font = normal
        ws.cell(row=i, column=9, value=f"=MIN(G{i},H{i})").font = neg
    nota = ws.cell(row=6, column=11, value=(
        f"Largura media de caractere em Arial = corpo x {R.MM_POR_PONTO} mm/pt "
        f"x {R.FATOR_CARACTERE} (fator para caixa mista). O teto editorial e "
        "decisao de Marketing: acima dele o nome deixa de ler na gondola, "
        "mesmo cabendo. MAX CARAC NOME e o menor dos dois."))
    nota.alignment = Alignment(wrap_text=True, vertical="top")
    nota.font = normal
    ws.column_dimensions["I"].width = 15
    ws.column_dimensions["K"].width = 60

    # --- SKUs por layout, com COUNTIF vivo sobre a Matriz_SKU
    r0 = 6 + len(R.LAYOUTS) + 2
    titulo(r0, "2. SKUS POR LAYOUT (conta a Matriz_SKU em tempo real)")
    for i, h in enumerate(["LAYOUT", "DESCRICAO", "SKUS", "% DO PORTFOLIO"], 1):
        c = ws.cell(row=r0 + 1, column=i, value=h)
        c.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=TURQUESA)
    desc = {1: "Pequeno - potes ate 400 ml", 2: "Medio - potes 400 ml a 1 L",
            3: "Grande - potes acima de 1 L e lixeiras",
            4: "Extra - organizadores e maletas"}
    fim = 3 + n_matriz
    for i, cod in enumerate(sorted(R.LAYOUTS), start=r0 + 2):
        ws.cell(row=i, column=1, value=cod).font = normal
        ws.cell(row=i, column=2, value=desc[cod]).font = normal
        ws.cell(row=i, column=3,
                value=f"=COUNTIF(Matriz_SKU!$D$4:$D${fim},A{i})").font = neg
        p = ws.cell(row=i, column=4,
                    value=f"=IFERROR(C{i}/SUM($C${r0+2}:$C${r0+1+len(R.LAYOUTS)}),0)")
        p.font = normal
        p.number_format = "0.0%"
    rt = r0 + 2 + len(R.LAYOUTS)
    ws.cell(row=rt, column=2, value="TOTAL").font = neg
    ws.cell(row=rt, column=3,
            value=f"=SUM(C{r0+2}:C{rt-1})").font = neg

    # --- proposta de selos por categoria
    r1 = rt + 2
    titulo(r1, "3. SELOS PROPOSTOS POR CATEGORIA (validar com Engenharia)")
    ws.cell(row=r1 + 1, column=1, value="CATEGORIA").font = neg
    ws.cell(row=r1 + 1, column=2, value="SELOS").font = neg
    linha = r1 + 2
    for cat, selos in R.SELOS_POR_CATEGORIA.items():
        ws.cell(row=linha, column=1, value=cat).font = normal
        ws.cell(row=linha, column=2,
                value=",".join(R.SELOS_UNIVERSAIS + selos)).font = normal
        linha += 1

    # --- proposta de aplicacoes por categoria
    linha += 1
    titulo(linha, "4. APLICACOES PROPOSTAS POR CATEGORIA (ordem = prioridade)")
    linha += 1
    ws.cell(row=linha, column=1, value="CATEGORIA").font = neg
    ws.cell(row=linha, column=2, value="APLICACOES").font = neg
    linha += 1
    for cat, aps in R.APLICACOES_POR_CATEGORIA.items():
        ws.cell(row=linha, column=1, value=cat).font = normal
        ws.cell(row=linha, column=2, value=",".join(aps)).font = normal
        linha += 1
    return ws


NOTA_INSTRUCOES = [
    ("", None),
    ("O QUE JA VEM PREENCHIDO NESTE FICHEIRO", True),
    ("A Matriz_SKU foi preenchida a partir da folha Produtos Nitron 2026, "
     "635 SKUs, por script. As colunas REFERENCIA e DESCRICAO SANKHYA sao "
     "copia exacta da origem. As restantes sao proposta e precisam de revisao.",
     False),
    ("CODPROD ficou vazio de proposito: o codigo interno do Sankhya nao consta "
     "da folha de produtos. Preencher por PROCV sobre a REFERENCIA antes de "
     "importar.", False),
    ("Separador Auditoria: uma linha por SKU com a razao de cada decisao e a "
     "coluna ALERTAS. Comecar a revisao por ai, filtrando ALERTAS nao vazio.",
     False),
    ("Separador Regras_Auto: os criterios usados e o calculo do MAX CARAC NOME. "
     "Mudar criterio e correr de novo o gerador, nao corrigir 635 linhas a mao.",
     False),
    ("Separador Impressao_Interna: o mesmo conteudo mais o EAN, no formato que "
     "o software da impressora de codigo de barras consome. O EAN nao entra na "
     "Matriz_SKU porque nao e conteudo grafico da bobina.", False),
    ("Os selos sao proposta por categoria mais evidencia no texto. Livre de BPA, "
     "apto para micro-ondas e apto para congelador sao declaracoes tecnicas: "
     "exigem assinatura da Engenharia/Qualidade antes de irem para chapa.",
     False),
    ("Foram acrescentadas ao Cat_Aplicacoes as aplicacoes que faltavam para "
     "cobrir o portfolio, marcadas com NOVO = S. Cada uma obriga a um icone novo.",
     False),
]


def acrescentar_nota_instrucoes(ws):
    linha = ws.max_row + 1
    for texto, negrito in NOTA_INSTRUCOES:
        c = ws.cell(row=linha, column=1, value=texto or None)
        if negrito is True:
            c.font = Font(name="Arial", size=10, bold=True)
        elif negrito is False:
            c.font = Font(name="Arial", size=10)
        c.alignment = Alignment(wrap_text=True)
        linha += 1


def exportar_csv(wb, nomes):
    DIR_CSV.mkdir(parents=True, exist_ok=True)
    for nome in nomes:
        ws = wb[nome]
        dados = [[c.value for c in row] for row in ws.iter_rows(min_row=3)]
        dados = [r for r in dados if any(v is not None for v in r)]
        df = pd.DataFrame(dados[1:], columns=dados[0])
        destino = DIR_CSV / f"{nome.lower()}.csv"
        df.to_csv(destino, sep=";", index=False, encoding="utf-8-sig")
        print(f"  csv  {destino.relative_to(RAIZ)}  ({len(df)} linhas)")


def validar_codigos(wb, linhas):
    """Nenhum codigo pode sair da matriz sem existir no catalogo."""
    def catalogo(aba):
        ws = wb[aba]
        return {ws.cell(row=r, column=1).value
                for r in range(4, ws.max_row + 1)
                if ws.cell(row=r, column=1).value}

    selos_ok, aplic_ok = catalogo("Cat_Selos"), catalogo("Cat_Aplicacoes")
    erros = []
    for l in linhas:
        for cod in (l[13] or "").split(","):
            if cod and cod not in selos_ok:
                erros.append(f"{l[1]}: selo {cod} nao existe no Cat_Selos")
        for cod in (l[14] or "").split(","):
            if cod and cod not in aplic_ok:
                erros.append(f"{l[1]}: aplicacao {cod} nao existe no Cat_Aplicacoes")
    if erros:
        for e in erros[:20]:
            print("  ERRO", e)
        raise SystemExit(f"{len(erros)} codigos fora do catalogo. Corrigir "
                         "scripts/etiquetas/regras.py antes de entregar.")
    print(f"  validacao: {len(selos_ok)} selos e {len(aplic_ok)} aplicacoes "
          "no catalogo, nenhum codigo orfao na matriz")


def ler_produtos(caminho: Path) -> pd.DataFrame:
    if caminho.suffix.lower() in (".xlsx", ".xlsm"):
        return pd.read_excel(caminho, dtype=str)
    return pd.read_csv(caminho, sep=";", dtype=str)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--produtos", type=Path, default=PRODUTOS)
    ap.add_argument("--modelo", type=Path, default=MODELO)
    ap.add_argument("--saida", type=Path, default=SAIDA)
    args = ap.parse_args()

    df = ler_produtos(args.produtos)
    df = df[df["REF"].notna()].copy()
    print(f"produtos lidos: {len(df)}")

    linhas, auditoria = construir(df)

    wb = load_workbook(args.modelo)
    preencher_matriz(wb["Matriz_SKU"], linhas)
    preencher_layouts(wb["Layouts"])
    novas = acrescentar_aplicacoes_novas(wb["Cat_Aplicacoes"])
    validar_codigos(wb, linhas)
    aba_auditoria(wb, auditoria)
    aba_impressao(wb, df, linhas)
    aba_regras(wb, linhas, len(linhas))
    acrescentar_nota_instrucoes(wb["Instrucoes"])

    # As formulas do separador Regras_Auto sao escritas sem valor em cache.
    # fullCalcOnLoad obriga o Excel e o LibreOffice a calcular ao abrir.
    wb.calculation.fullCalcOnLoad = True
    args.saida.parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.saida)
    print(f"escrito: {args.saida.relative_to(RAIZ)}")
    print(f"  matriz: {len(linhas)} SKUs | aplicacoes novas: {novas}")

    exportar_csv(load_workbook(args.saida),
                 ["Matriz_SKU", "Cat_Selos", "Cat_Aplicacoes", "Layouts",
                  "Unidades", "Impressao_Interna", "Auditoria"])

    com_alerta = sum(1 for a in auditoria if a[-1])
    print(f"  SKUs com alerta na Auditoria: {com_alerta}")


if __name__ == "__main__":
    main()
